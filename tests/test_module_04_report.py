from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPORT_FILE = PROJECT_ROOT / "reports" / "financial_report.xlsx"


def main():
    assert REPORT_FILE.exists(), f"Relatorio nao encontrado: {REPORT_FILE}"

    workbook = load_workbook(REPORT_FILE, read_only=True, data_only=True)
    expected_sheets = {
        "Resumo Executivo",
        "Selic Diaria",
        "IPCA Mensal",
        "Cotacoes B3",
        "Benchmarks",
        "Cobertura",
        "Performance",
    }

    assert expected_sheets.issubset(set(workbook.sheetnames))

    print("\nTeste do Modulo 4 concluido com sucesso.")


if __name__ == "__main__":
    main()
