from __future__ import annotations
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FONT = Font(color="1F4E78", bold=True, size=14)
SECTION_FILL = PatternFill("solid", fgColor="E2F0D9")
SECTION_FONT = Font(color="375623", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def _read_sql(database_file, query):
    if not database_file.exists():
        raise FileNotFoundError(f"Banco nao encontrado: {database_file}")
    with sqlite3.connect(database_file) as conn:
        return pd.read_sql_query(query, conn)

def load_report_data(database_file):
    selic_df = _read_sql(database_file, "SELECT reference_date, value AS selic_daily_value FROM vw_bcb_series_values WHERE series_name = 'selic_daily' ORDER BY reference_date")
    ipca_df = _read_sql(database_file, "SELECT reference_date, value AS ipca_monthly_value FROM vw_bcb_series_values WHERE series_name = 'ipca_monthly' ORDER BY reference_date")
    bcb_series_df = _read_sql(database_file, "SELECT series_name, description, frequency, reference_date, value FROM vw_bcb_series_values ORDER BY series_name, reference_date")
    bcb_snapshot_df = _read_sql(database_file, "SELECT * FROM vw_bcb_latest_snapshot ORDER BY series_name")
    b3_returns_df = _read_sql(database_file, "SELECT * FROM vw_b3_asset_returns ORDER BY asset_type, ticker")
    stocks_df = _read_sql(database_file, "SELECT reference_date, ticker, asset_type, open_price, high_price, low_price, close_price, adjusted_close_price, volume FROM vw_b3_stock_prices ORDER BY reference_date, ticker")
    selic_df["reference_date"] = pd.to_datetime(selic_df["reference_date"])
    ipca_df["reference_date"] = pd.to_datetime(ipca_df["reference_date"])
    bcb_series_df["reference_date"] = pd.to_datetime(bcb_series_df["reference_date"])
    bcb_snapshot_df["last_available_date"] = pd.to_datetime(bcb_snapshot_df["last_available_date"])
    b3_returns_df["start_date"] = pd.to_datetime(b3_returns_df["start_date"])
    b3_returns_df["end_date"] = pd.to_datetime(b3_returns_df["end_date"])
    stocks_df["reference_date"] = pd.to_datetime(stocks_df["reference_date"])
    return {
        "selic": selic_df,
        "ipca": ipca_df,
        "bcb_series": bcb_series_df,
        "bcb_snapshot": bcb_snapshot_df,
        "b3_returns": b3_returns_df,
        "stocks": stocks_df,
    }

def build_executive_metrics(data):
    selic_df = data["selic"].copy()
    ipca_df = data["ipca"].copy()
    stocks_df = data["stocks"].copy()
    latest = selic_df.sort_values("reference_date").iloc[-1]
    latest_selic = {"date": latest["reference_date"], "value": float(latest["selic_daily_value"])}
    bcb_snapshot = data["bcb_snapshot"].set_index("series_name")
    ipca_2024 = ipca_df[ipca_df["reference_date"].dt.year == 2024]
    ipca_acc = ((ipca_2024["ipca_monthly_value"] / 100 + 1).prod() - 1) * 100 if not ipca_2024.empty else None
    stock_returns = []
    for ticker, tdf in stocks_df.groupby("ticker"):
        tdf = tdf.sort_values("reference_date")
        first_price = float(tdf.iloc[0]["adjusted_close_price"])
        last_price = float(tdf.iloc[-1]["adjusted_close_price"])
        return_pct = ((last_price / first_price) - 1) * 100 if first_price != 0 else None
        stock_returns.append({"ticker": ticker, "start_date": tdf.iloc[0]["reference_date"], "end_date": tdf.iloc[-1]["reference_date"], "first_adjusted_close": first_price, "last_adjusted_close": last_price, "return_pct": return_pct})
    return {
        "latest_selic": latest_selic,
        "latest_usd_ptax": bcb_snapshot.loc["usd_brl_ptax_sell_daily"].to_dict() if "usd_brl_ptax_sell_daily" in bcb_snapshot.index else None,
        "latest_cdi": bcb_snapshot.loc["cdi_daily"].to_dict() if "cdi_daily" in bcb_snapshot.index else None,
        "ipca_accumulated_2024": ipca_acc,
        "stock_returns": stock_returns,
    }

def _write_dataframe(worksheet, df, start_row, start_col, table_name=None):
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.date
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell = worksheet.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
    last_row = start_row + len(df)
    last_col = start_col + len(df.columns) - 1
    if table_name:
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        worksheet.add_table(table)
    return last_row, last_col

def _format_worksheet_columns(worksheet, max_width=24):
    for col_cells in worksheet.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def _fmt_date(ws, col, start, end):
    for r in range(start, end + 1):
        ws[f"{col}{r}"].number_format = "dd/mm/yyyy"

def _fmt_num(ws, col, start, end, fmt):
    for r in range(start, end + 1):
        ws[f"{col}{r}"].number_format = fmt

def create_summary_sheet(workbook, metrics, data):
    ws = workbook.active
    ws.title = "Resumo Executivo"
    ws["A1"] = "Brazilian Financial Data Pipeline"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="1F4E78", bold=True, size=16)
    ws["A2"] = "Resumo Executivo"
    ws["A2"].font = TITLE_FONT
    ws["A4"] = "Gerado em"
    ws["B4"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws["A7"] = "Metricas principais"
    ws["A7"].fill = SECTION_FILL
    ws["A7"].font = SECTION_FONT
    ws["A9"] = "Ultima Selic diaria"
    ws["B9"] = metrics["latest_selic"]["value"]
    ws["C9"] = metrics["latest_selic"]["date"].date()
    ws["A10"] = "IPCA acumulado 2024"
    ws["B10"] = (metrics["ipca_accumulated_2024"] / 100) if metrics["ipca_accumulated_2024"] else None
    ws["A11"] = "Dolar PTAX venda"
    ws["B11"] = metrics["latest_usd_ptax"]["latest_value"] if metrics["latest_usd_ptax"] else None
    ws["C11"] = metrics["latest_usd_ptax"]["last_available_date"].date() if metrics["latest_usd_ptax"] else None
    ws["A12"] = "CDI diario"
    ws["B12"] = metrics["latest_cdi"]["latest_value"] if metrics["latest_cdi"] else None
    ws["C12"] = metrics["latest_cdi"]["last_available_date"].date() if metrics["latest_cdi"] else None
    ws["B9"].number_format = "0.000000"
    ws["C9"].number_format = "dd/mm/yyyy"
    ws["B10"].number_format = "0.00%"
    ws["B11"].number_format = "R$ #,##0.0000"
    ws["C11"].number_format = "dd/mm/yyyy"
    ws["B12"].number_format = "0.000000"
    ws["C12"].number_format = "dd/mm/yyyy"
    ws["H7"] = "Volume de dados"
    ws["H7"].fill = SECTION_FILL
    ws["H7"].font = SECTION_FONT
    ws["H9"] = "Registros Selic"
    ws["I9"] = len(data["selic"])
    ws["H10"] = "Registros IPCA"
    ws["I10"] = len(data["ipca"])
    ws["H11"] = "Registros B3"
    ws["I11"] = len(data["stocks"])
    ws["H12"] = "Tickers"
    ws["I12"] = data["stocks"]["ticker"].nunique()
    ws["H13"] = "Series BCB"
    ws["I13"] = data["bcb_series"]["series_name"].nunique()
    returns_df = pd.DataFrame(metrics["stock_returns"])[["ticker", "start_date", "end_date", "first_adjusted_close", "last_adjusted_close", "return_pct"]]
    returns_df["return_pct"] = returns_df["return_pct"] / 100
    ws["A15"] = "Retorno das acoes no periodo"
    ws["A15"].fill = SECTION_FILL
    ws["A15"].font = SECTION_FONT
    last_row, _ = _write_dataframe(ws, returns_df, 17, 1, "tbl_summary_stock_returns")
    for r in range(18, last_row + 1):
        ws[f"B{r}"].number_format = "dd/mm/yyyy"
        ws[f"C{r}"].number_format = "dd/mm/yyyy"
        ws[f"D{r}"].number_format = "R$ #,##0.00"
        ws[f"E{r}"].number_format = "R$ #,##0.00"
        ws[f"F{r}"].number_format = "0.00%"
    _format_worksheet_columns(ws, 38)

def create_selic_sheet(workbook, selic_df):
    ws = workbook.create_sheet("Selic Diaria")
    ws["A1"] = "Selic Diaria"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    df = selic_df.rename(columns={"reference_date": "data", "selic_daily_value": "selic_diaria"})
    last_row, _ = _write_dataframe(ws, df, 3, 1, "tbl_selic_daily")
    _fmt_date(ws, "A", 4, last_row)
    _fmt_num(ws, "B", 4, last_row, "0.000000")
    chart = LineChart()
    chart.title = "Selic diaria"
    chart.height = 10
    chart.width = 22
    chart.add_data(Reference(ws, min_col=2, min_row=3, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=last_row))
    ws.add_chart(chart, "D3")
    ws.freeze_panes = "A4"
    _format_worksheet_columns(ws, 18)

def create_ipca_sheet(workbook, ipca_df):
    ws = workbook.create_sheet("IPCA Mensal")
    ws["A1"] = "IPCA Mensal"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    df = ipca_df.rename(columns={"reference_date": "data", "ipca_monthly_value": "ipca_mensal"})
    last_row, _ = _write_dataframe(ws, df, 3, 1, "tbl_ipca_monthly")
    _fmt_date(ws, "A", 4, last_row)
    _fmt_num(ws, "B", 4, last_row, "0.00")
    chart = BarChart()
    chart.title = "IPCA mensal"
    chart.height = 10
    chart.width = 22
    chart.add_data(Reference(ws, min_col=2, min_row=3, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=last_row))
    ws.add_chart(chart, "D3")
    ws.freeze_panes = "A4"
    _format_worksheet_columns(ws, 18)

def create_stocks_sheet(workbook, stocks_df):
    ws = workbook.create_sheet("Cotacoes B3")
    ws["A1"] = "Cotacoes B3"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    df = stocks_df.rename(columns={"reference_date": "data", "open_price": "abertura", "high_price": "maxima", "low_price": "minima", "close_price": "fechamento", "adjusted_close_price": "fechamento_ajustado", "volume": "volume"})
    last_row, _ = _write_dataframe(ws, df, 3, 1, "tbl_b3_quotes")
    _fmt_date(ws, "A", 4, last_row)
    for col in ["D", "E", "F", "G", "H"]:
        _fmt_num(ws, col, 4, last_row, "R$ #,##0.00")
    _fmt_num(ws, "I", 4, last_row, "#,##0")
    pivot_df = stocks_df.pivot_table(index="reference_date", columns="ticker", values="close_price", aggfunc="last").reset_index().sort_values("reference_date")
    pivot_df = pivot_df.rename(columns={"reference_date": "data"})
    pivot_last_row, pivot_last_col = _write_dataframe(ws, pivot_df, 3, 10, "tbl_b3_chart_prices")
    chart = LineChart()
    chart.title = "Fechamento diario - ativos e benchmarks"
    chart.height = 12
    chart.width = 26
    chart.add_data(Reference(ws, min_col=11, max_col=pivot_last_col, min_row=3, max_row=pivot_last_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=10, min_row=4, max_row=pivot_last_row))
    ws.add_chart(chart, "J20")
    ws.freeze_panes = "A4"
    _format_worksheet_columns(ws, 22)


def create_benchmarks_sheet(workbook, data):
    ws = workbook.create_sheet("Benchmarks")
    ws["A1"] = "Benchmarks e Macro"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Snapshot BCB"
    ws["A3"].fill = SECTION_FILL
    ws["A3"].font = SECTION_FONT
    bcb_snapshot_df = data["bcb_snapshot"].rename(columns={
        "series_name": "serie",
        "description": "descricao",
        "frequency": "frequencia",
        "last_available_date": "ultima_data",
        "latest_value": "ultimo_valor",
    })
    snapshot_last_row, _ = _write_dataframe(ws, bcb_snapshot_df, 5, 1, "tbl_bcb_snapshot")
    _fmt_date(ws, "D", 6, snapshot_last_row)
    _fmt_num(ws, "E", 6, snapshot_last_row, "0.000000")

    ws["H3"] = "Retorno B3 no periodo"
    ws["H3"].fill = SECTION_FILL
    ws["H3"].font = SECTION_FONT
    returns_df = data["b3_returns"].copy()
    returns_df["return_pct"] = returns_df["return_pct"] / 100
    returns_last_row, _ = _write_dataframe(ws, returns_df, 5, 8, "tbl_b3_returns")
    for r in range(6, returns_last_row + 1):
        ws[f"J{r}"].number_format = "dd/mm/yyyy"
        ws[f"K{r}"].number_format = "dd/mm/yyyy"
        ws[f"L{r}"].number_format = "R$ #,##0.00"
        ws[f"M{r}"].number_format = "R$ #,##0.00"
        ws[f"N{r}"].number_format = "0.00%"

    _format_worksheet_columns(ws, 28)

def create_financial_report(database_file, output_file):
    output_file.parent.mkdir(parents=True, exist_ok=True)
    data = load_report_data(database_file)
    metrics = build_executive_metrics(data)
    wb = Workbook()
    create_summary_sheet(wb, metrics, data)
    create_selic_sheet(wb, data["selic"])
    create_ipca_sheet(wb, data["ipca"])
    create_stocks_sheet(wb, data["stocks"])
    create_benchmarks_sheet(wb, data)
    wb.save(output_file)
    return {
        "output_file": str(output_file),
        "selic_rows": len(data["selic"]),
        "ipca_rows": len(data["ipca"]),
        "stock_rows": len(data["stocks"]),
        "bcb_series_rows": len(data["bcb_series"]),
        "tickers": sorted(data["stocks"]["ticker"].unique().tolist()),
        "latest_selic_date": metrics["latest_selic"]["date"].strftime("%Y-%m-%d"),
        "latest_selic_value": metrics["latest_selic"]["value"],
        "ipca_accumulated_2024": metrics["ipca_accumulated_2024"],
    }
